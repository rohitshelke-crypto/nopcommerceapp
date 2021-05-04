
##we are performing data  driven testing on login test case.
# first we need to add excel file in our created TestData folder,we can directly copy that excel file and paste it in our TestData foldeer
##In same way we can copy and paste xlutility file and paste it in our utility folder


import openpyxl



def getRowcount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_row)

def getColoumnCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_Coloumn)

def readData(file,sheetname,rownum,coloumnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum,column=coloumnno).value

def writeData(file,sheetname,rownum,coloumnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum,column=coloumnno).value=data
    workbook.save(file)


